import openpyxl
import psycopg

# Establish a connection to the PostgreSQL database
conn = psycopg.connect(
    host="127.0.0.1",
    port="5432",
    dbname="moviedb",
    user="postgres",
    password=""
)

# Create a cursor object to interact with the database
cursor = conn.cursor()

USERS=[]
AUDIENCES=[]
DIRECTORS=[]
MOVIE_SESSIONS=[]
GENRES=[]
RATING_PLATFORMS=[]
RATINGS=[]
DB_MANAGERS=[]

THEATERS=[]
THEATER_IDS=[]

MOVIES=[] #INCLUDE genre_list, predecessors_list
MOVIE_IDS=[]



def read_excel_file(filename):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename)

    # Get the sheet names
    sheet_names = workbook.sheetnames

    # Iterate through each sheet
    for sheet_name in sheet_names:
        # Select the sheet by name
        sheet = workbook[sheet_name]
        
        if sheet_name == "User":
            for row in sheet.iter_rows(values_only=True):
                username, password, name, surname = row[0], row[1], row[2], row[3]
                if username == "username" or username == None: 
                    continue
                values = [username, password, name, surname]
                USERS.append(values)

        elif sheet_name == "Audience":
            for row in sheet.iter_rows(values_only=True):
                username, list_of_bought_sessions, list_of_rating_platforms = row[0], str(row[1]).split(","), str(row[2]).split(",")
                if username == "username" or username == None: 
                    continue
                values = [username, list_of_bought_sessions, list_of_rating_platforms]
                AUDIENCES.append(values)

        elif sheet_name == "Director":
            for row in sheet.iter_rows(values_only=True):
                username, nationality, platform_id = row[0], row[1], row[2]
                if username == "username" or username == None: 
                    continue
                values = [username, nationality, platform_id]
                DIRECTORS.append(values)
        elif sheet_name == "Movie_Sessions":
            for row in sheet.iter_rows(values_only=True):
                session_id, movie_id, name, duration, genre_list, overall_rating, director_username, director_platform_id, predecessors_list, theatre_id, theatre_name, theatre_capacity, district, time_slot, date = row[0],row[1],row[2],row[3],str(row[4]).split(","),row[5],row[6],row[7], str(row[8]).split(","),row[9],row[10],row[11],row[12],row[13],row[14]
                if session_id == "session_id" or session_id == None:
                    continue
                if theatre_id not in THEATER_IDS:
                    THEATER_IDS.append(theatre_id)
                    THEATERS.append([theatre_id,theatre_name,theatre_capacity,district])
                if movie_id not in MOVIE_IDS:
                    MOVIE_IDS.append(movie_id)
                    MOVIES.append([movie_id,name,duration,overall_rating,genre_list,predecessors_list,director_username])
                values = [session_id, movie_id, name, duration, genre_list, overall_rating, director_username, director_platform_id, predecessors_list, theatre_id, theatre_name, theatre_capacity, district, time_slot, date]
                MOVIE_SESSIONS.append(values)
        elif sheet_name == "Genre":
            for row in sheet.iter_rows(values_only=True):
                genre_id, genre_name = row[0], row[1]
                if genre_id == "genre_id" or genre_id == None:
                    continue
                values = [genre_id, genre_name]
                GENRES.append(values)
        elif sheet_name == "Rating_Platform":
            for row in sheet.iter_rows(values_only=True):
                platform_id, platform_name = row[0], row[1]
                if platform_id == "platform_id" or platform_id == None:
                    continue
                values = [platform_id, platform_name]
                RATING_PLATFORMS.append(values)
        elif sheet_name == "Rating":
            for row in sheet.iter_rows(values_only=True):
                username, movie_id, rating = row[0],row[1],row[2]
                if username == "username" or username == None:
                    continue
                values = [username, movie_id, rating]
                RATINGS.append(values)
        elif sheet_name == "Database_Manager":
            for row in sheet.iter_rows(values_only=True):
                username, password = row[0], row[1]
                if username == "username" or username == None:
                    continue
                values = [username, password]
                DB_MANAGERS.append(values)
            
    # Close the workbook
    workbook.close()

def print_data():
    print( USERS)
    print(30 * "-")
    print( AUDIENCES)
    print(30 * "-")
    print( DIRECTORS)
    print(30 * "-")
    print( MOVIE_SESSIONS)
    print(30 * "-")
    print( GENRES)
    print(30 * "-")
    print( RATING_PLATFORMS)
    print(30 * "-")
    print( RATINGS)
    print(30 * "-")
    print( DB_MANAGERS)

def insert_db_managers():
    manager_number=0 #count of managers
    for row in DB_MANAGERS:
        username, password = row[0], row[1]
        manager_number += 1
        query = "INSERT INTO database_manager (username, manager_number, password1) VALUES (%s, %s, %s);"
        values = (username, manager_number, password)
        cursor.execute(query, values)
        conn.commit()
def delete_db_managers():
    query = "TRUNCATE TABLE database_manager CASCADE;"
    cursor.execute(query)
    conn.commit()

def insert_rating_platforms():
    for row in RATING_PLATFORMS:
        platform_id, platform_name = row[0], row[1]
        query = "INSERT INTO rating_platform (platform_id, platform_name) VALUES (%s, %s);"
        values = (platform_id, platform_name)
        cursor.execute(query, values)
        conn.commit()

def delete_rating_platforms():
    query = "TRUNCATE TABLE rating_platform CASCADE;"
    cursor.execute(query)
    conn.commit()

def insert_genres():
    for row in GENRES:
        genre_id, genre_name = row[0], row[1]
        query = "INSERT INTO genre (genre_id, genre_name) VALUES (%s, %s);"
        values = (genre_id, genre_name)
        cursor.execute(query, values)
        conn.commit()
def delete_genres():
    query = "TRUNCATE TABLE genre CASCADE;"
    cursor.execute(query)
    conn.commit()

def insert_movies():
    for row in MOVIES:
        movie_id,name,duration,overall_rating,genre_list,predecessors_list, director_username = row[0], row[1], row[2], row[3], row[4], row[5], row[6]
    
        query = "INSERT INTO movie (movie_id,movie_name,duration,overall_rating) VALUES (%s, %s, %s, %s);"
        values = (movie_id,name,duration,overall_rating)
        cursor.execute(query, values)
        conn.commit()
        for genre_id in genre_list:
            if genre_id == None:
                continue
            query = "INSERT INTO genre_list (genre_id,movie_id) VALUES (%s, %s);"
            values = (genre_id, movie_id)
            cursor.execute(query, values)
            conn.commit()
        query = "INSERT INTO directs (username, movie_id) VALUES (%s, %s);"
        values = (director_username, movie_id)
        cursor.execute(query, values)
        conn.commit()


def delete_movies():
    query = "TRUNCATE TABLE movie CASCADE;"
    cursor.execute(query)
    conn.commit()
    query = "TRUNCATE TABLE genre_list CASCADE;"
    cursor.execute(query)
    conn.commit()
    query = "TRUNCATE TABLE directs CASCADE;"
    cursor.execute(query)
    conn.commit()

def insert_predecessors():
    for row in MOVIES:
        movie_id,name,duration,overall_rating,genre_list,predecessors_list = row[0], row[1], row[2], row[3], row[4], row[5]
        for predecessors_id in predecessors_list:
            if predecessors_id == None or predecessors_id == 'None':
                continue
            query = "INSERT INTO predecessors (predecessor_movie_id,succeeds_movie_id) VALUES (%s, %s);"
            values = (predecessors_id, movie_id)
            cursor.execute(query, values)
            conn.commit()

def delete_predecessors():
    query = "TRUNCATE TABLE predecessors CASCADE;"
    cursor.execute(query)
    conn.commit()

def insert_theaters():
    for row in THEATERS:
        theatre_id,theatre_name,theatre_capacity,district = row[0], row[1], row[2], row[3]
        query = "INSERT INTO theater (theater_id,theater_name,theater_capacity,district) VALUES (%s, %s, %s, %s);"
        values = (theatre_id,theatre_name,theatre_capacity,district)
        cursor.execute(query, values)
        conn.commit()

def delete_theaters():
    query = "TRUNCATE TABLE theater CASCADE;"
    cursor.execute(query)
    conn.commit()

def insert_ratings():
    rating_count = 0 #Count determines id 
    for row in RATINGS:
        username, movie_id, rating = row[0],row[1],row[2]
        values = (rating_count, username, movie_id, rating)
        rating_count += 1
        query = "INSERT INTO ratings (rating_id, username, movie_id,rating) VALUES (%s, %s, %s, %s);"
        cursor.execute(query, values)
        conn.commit()
def delete_ratings():
    query = "TRUNCATE TABLE ratings CASCADE;"
    cursor.execute(query)
    conn.commit()

def insert_users():

    for row in USERS:
        username, password, name, surname = row[0], row[1], row[2], row[3]
        values = (username, password, name, surname)
        query = "INSERT INTO user1 (username, password1, surname, name1) VALUES (%s, %s, %s, %s);"
        cursor.execute(query, values)
        conn.commit()
    
    conn.commit()

def delete_users():
    query = "TRUNCATE TABLE user1 CASCADE;"
    cursor.execute(query)
    conn.commit()

def insert_directors():
    for row in DIRECTORS:
        username, nationality, platform_id = row[0], row[1], row[2]
        values = (username, nationality)
        query = "INSERT INTO director (username, nationality) VALUES (%s, %s);"
        cursor.execute(query, values)
        conn.commit()

        values = (username, platform_id)
        query = "INSERT INTO registered (username, platform_id) VALUES (%s, %s);"
        cursor.execute(query, values)
        conn.commit()

def delete_directors():
    query = "TRUNCATE TABLE director CASCADE;"
    cursor.execute(query)
    conn.commit()
    query = "TRUNCATE TABLE registered CASCADE;"
    cursor.execute(query)
    conn.commit()

def insert_audience():
    for row in AUDIENCES:
        username, list_of_bought_sessions, list_of_rating_platforms = row[0], row[1], row[2]
        values = (username)
        query = "INSERT INTO audience (username) VALUES ('{0}');".format(username)
        cursor.execute(query)
        conn.commit()

        query = "INSERT INTO bought_tickets (username, session_id) VALUES (%s, %s);"
        for session in list_of_bought_sessions:
            values = (username,session)
            cursor.execute(query, values)
            conn.commit()

        query = "INSERT INTO subscriptions (username, platform_id) VALUES (%s, %s);"
        for platform_id in list_of_rating_platforms:
            values = (username, platform_id)
            cursor.execute(query, values)
            conn.commit()

def delete_audience():
    query = "TRUNCATE TABLE audience CASCADE;"
    cursor.execute(query)
    conn.commit()
    query = "TRUNCATE TABLE bought_tickets CASCADE;"
    cursor.execute(query)
    conn.commit()
    query = "TRUNCATE TABLE subscriptions CASCADE;"
    cursor.execute(query)
    conn.commit()

def insert_movie_sessions():
    occupience_count=0
    for row in MOVIE_SESSIONS:
        session_id, movie_id, name, duration, genre_list, overall_rating, director_username, director_platform_id, predecessors_list, theatre_id, theatre_name, theatre_capacity, district, time_slot, date = row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7], row[8],row[9],row[10],row[11],row[12],row[13],row[14]
        query = "INSERT INTO movie_session (session_id) VALUES ('{0}');".format(session_id)
        cursor.execute(query)
        conn.commit()
        for i in range(duration):
            values = (occupience_count,date,int(time_slot)+i,theatre_id)
            query = "INSERT INTO occupience_info (occupience_id , time1, slot_number, theater_id) VALUES (%s, %s, %s, %s);"
            cursor.execute(query,values)
            conn.commit()
            values= (session_id,occupience_count)
            query = "INSERT INTO place (session_id, occupience_id) VALUES ( %s, %s);"
            cursor.execute(query,values)
            conn.commit()
            occupience_count += 1
        values= (session_id,movie_id)
        query = "INSERT INTO has_movie (session_id, movie_id) VALUES ( %s, %s);"
        cursor.execute(query,values)
        conn.commit()

def delete_movie_sessions():
    query = "TRUNCATE TABLE movie_session CASCADE;"
    cursor.execute(query)
    conn.commit()

    query = "TRUNCATE TABLE occupience_info CASCADE;"
    cursor.execute(query)
    conn.commit()

    query = "TRUNCATE TABLE place CASCADE;"
    cursor.execute(query)
    conn.commit()
    query = "TRUNCATE TABLE has_movie CASCADE;"
    cursor.execute(query)
    conn.commit()

def delete_all():
    delete_audience()
    delete_db_managers()
    delete_directors()
    delete_genres()
    delete_movie_sessions()
    delete_movies()
    delete_predecessors()
    delete_rating_platforms()
    delete_ratings()
    delete_theaters()
    delete_users()

def create_all():
    insert_db_managers()
    insert_rating_platforms()
    insert_users()
    insert_directors()
    insert_genres()
    insert_movies()
    insert_predecessors()
    insert_theaters()
    insert_movie_sessions()
    insert_audience()
    insert_ratings()




filename = 'Sample_Data.xlsx'
read_excel_file(filename)
delete_all()
create_all()


